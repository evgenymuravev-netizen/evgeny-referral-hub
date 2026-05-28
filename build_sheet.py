"""
Build a polished xlsx outreach sheet for Andrey, ready to upload to Google Drive
where it opens in Sheets. Contains:
  - Outreach Plan: 23 contacts with pre-drafted messages, hyperlinks, status dropdowns
  - Quick Reference: Evgeny's pitch + canonical rules
  - Templates: All 6 ask-type variants for reference
  - Needs Input: the 10 items where Andrey must identify the specific person
"""
import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


HERE = pathlib.Path(__file__).parent
DATA = json.loads((HERE / "active_outreach.json").read_text())
ITEMS = DATA["items"]

LINKEDIN_EVGENY = "https://linkedin.com/in/evgenymuravev"

# ─── Styling primitives ───────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A2238")
HEADER_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
BLOCK_FILL = PatternFill("solid", fgColor="7C5CFF")
BLOCK_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
LABEL_FILL = PatternFill("solid", fgColor="F5F7FF")
LABEL_FONT = Font(name="Inter", size=10, bold=True, color="1A2238")
BODY_FONT = Font(name="Inter", size=10, color="1A2238")
MONO_FONT = Font(name="JetBrains Mono", size=9, color="3A4566")
LINK_FONT = Font(name="Inter", size=10, color="0563C1", underline="single")
ACCENT_FILL = PatternFill("solid", fgColor="E8FCF6")
WARN_FILL = PatternFill("solid", fgColor="FFF4E0")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDE5"),
    right=Side(style="thin", color="DDDDE5"),
    top=Side(style="thin", color="DDDDE5"),
    bottom=Side(style="thin", color="DDDDE5"),
)


def first_name(name: str) -> str:
    """Extract first name, but keep 'Dr. Sebastian' or honorific intact."""
    if name.startswith("Dr. ") or name.startswith("Dr."):
        parts = name.split(" ", 2)
        return parts[1] if len(parts) > 1 else name
    return name.split(" ")[0]


def ask_area_from_detail(detail: str, fallback: str = "this area") -> str:
    """Pull the topic out of ask_detail like 'Ask if he has roles in digital transformation...'."""
    if "digital transformation" in detail.lower():
        return "digital transformation"
    if "openings" in detail.lower():
        return "product"
    return fallback


def generate_message(item):
    """Generate the primary pre-drafted message for this contact + ask_type."""
    name = first_name(item["name"])
    if item["name"].startswith("Dr."):
        name = item["name"]  # use full "Dr. Walter" for honorifics
    company = item.get("company") or "your company"
    ask = item["ask_type"]
    job_urls = item.get("job_urls") or []
    relocation = item.get("relocation")
    salary_note = item.get("salary_note")
    needs_input = item["linkedin"] is None
    # 2nd-degree without message_available → connection request short form
    # needs_input rows default to long-form DM; Andrey trims if connection req
    use_short = (
        item.get("degree") == 2
        and not item.get("message_available")
        and not needs_input
    )

    PITCH = "Evgeny helped us build our merchant lending and payments product, wallets for business in UAE"

    if ask == "post_in_chat":
        return (
            f"sharing the profile of Evgeny Muravev — he helped us build "
            f"Tabby's merchant lending and payments product, wallets for "
            f"business in UAE. Exploring new product roles in payments / "
            f"lending, UAE based, open to relocation.\n\n"
            f"LinkedIn: {LINKEDIN_EVGENY}\nCV attached.\n\n"
            f"happy to make any intro — DM me directly.\n\n— Andrey"
        )

    if ask == "send_cv":
        msg = (
            f"{name} — sending you the profile of Evgeny who is exploring "
            f"new product roles. {PITCH}. Would love it if you could keep "
            f"him in mind for any product roles in payments / lending"
        )
        if item.get("company"):
            msg += f" at {company}"
        msg += (
            f".\n\nhis LinkedIn: {LINKEDIN_EVGENY}\n"
            f"CV attached.\n\n— Andrey"
        )
        return msg

    if ask == "recommend_for_role":
        if use_short and not job_urls:
            # Short connection-request form
            return (
                f"{name} — would like to introduce you to Evgeny who built "
                f"Tabby's merchant lending, payments, and business wallets "
                f"in UAE. He's interested in product roles at {company} — "
                f"would love a quick chat. — Andrey"
            )
        msg = (
            f"{name} — would like to introduce you to Evgeny who is "
            f"exploring open roles at {company} in product. {PITCH} and "
            f"I highly recommend him.\n\n"
            f"He's specifically interested in these openings:\n"
        )
        for url in job_urls:
            msg += f"- {url}\n"
        if relocation:
            msg += f"\nHe's ready to relocate ({relocation}).\n"
        msg += (
            f"\nhis LinkedIn: {LINKEDIN_EVGENY}\n"
            f"happy to share his CV directly.\n\n— Andrey"
        )
        return msg

    if ask == "ask_about_openings":
        area = ask_area_from_detail(item["ask_detail"])
        msg = (
            f"{name} — hope you're doing well at {company}. Quick "
            f"question: are you guys hiring on {area}? If yes, I'd like to "
            f"introduce Evgeny who is exploring open roles at {company} "
            f"in product. {PITCH} — highly recommend.\n\n"
        )
        if job_urls:
            msg += "He's specifically interested in:\n"
            for url in job_urls:
                msg += f"- {url}\n"
            msg += "\n"
        msg += (
            f"his LinkedIn: {LINKEDIN_EVGENY}\n"
            f"happy to share his CV directly.\n\n— Andrey"
        )
        return msg

    if ask == "intro_to_founders":
        context = (
            item.get("ask_detail", "")
            .replace("Andrey previously offered to intro Evgeny to", "I previously offered to intro Evgeny to")
            .replace("Confirm and make the intro.", "")
            .strip()
        )
        andrey_note = item.get("andrey_notes", "")
        msg = (
            f"Hi — would like to introduce you to Evgeny who is exploring "
            f"new product opportunities. {PITCH}.\n\n"
            f"{context}\n"
        )
        if andrey_note:
            msg += f"\n{andrey_note}\n"
        msg += (
            f"\nIf you're open to a 20-min chat, I'll connect you both. "
            f"His LinkedIn: {LINKEDIN_EVGENY}\n\n— Andrey"
        )
        return msg

    if ask == "recommend_generally":
        if use_short:
            msg = (
                f"{name} — would like to introduce you to Evgeny who is "
                f"exploring open roles at {company} in product. {PITCH} "
                f"and I highly recommend him for any product role in "
                f"payments or lending."
            )
            if relocation:
                msg += f" Ready to relocate ({relocation})."
            return msg
        msg = (
            f"{name} — would like to introduce you to Evgeny who is "
            f"exploring open roles at {company} in product. {PITCH} and "
            f"I highly recommend him for any product role you may have "
            f"in payments or lending."
        )
        if relocation:
            msg += f" Ready to relocate ({relocation})."
        if salary_note:
            msg += f" {salary_note}."
        msg += (
            f"\n\nhis LinkedIn: {LINKEDIN_EVGENY}\n"
            f"happy to share his CV directly.\n\n— Andrey"
        )
        return msg

    return f"(no template for ask_type={ask})"


def char_limit(item) -> int:
    """The max chars for this contact's primary message."""
    needs_input = item["linkedin"] is None
    # Connection-request 300-char form ONLY when we KNOW the contact is
    # 2nd-degree with no DM access. needs_input items default to DM limit
    # since Andrey will decide channel after identifying the person.
    use_short = (
        item.get("degree") == 2
        and not item.get("message_available")
        and not needs_input
    )
    if item["ask_type"] == "post_in_chat":
        return 1600  # WhatsApp/Telegram chat
    return 300 if use_short else 1900


# ─── Workbook ─────────────────────────────────────────────────────────────
wb = Workbook()

# === SHEET 1: Outreach Plan ===
ws = wb.active
ws.title = "Outreach Plan"

HEADERS = [
    "#",            # priority
    "ID",
    "Block",
    "Name",
    "Role",
    "Company",
    "Ask Type",
    "LinkedIn",
    "Job URLs (if any)",
    "Relocation",
    "Pre-drafted message",
    "Chars",
    "Limit",
    "Channel",
    "Status",
    "Sent date",
    "Reply?",
    "Notes",
]

for col, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Sort by priority
items_sorted = sorted(ITEMS, key=lambda x: x["priority"])

row = 2
for item in items_sorted:
    message = generate_message(item)
    chars = len(message)
    limit = char_limit(item)
    job_urls_str = "\n".join(item.get("job_urls") or [])

    cells = [
        item["priority"],
        item["id"],
        item["block"],
        item["name"],
        item.get("role") or "",
        item.get("company") or "",
        item["ask_type"],
        item.get("linkedin") or "(needs Andrey input)",
        job_urls_str,
        item.get("relocation") or "",
        message,
        chars,
        limit,
        item.get("channel") or "LinkedIn",
        "Not started",
        "",  # sent date
        "",  # reply
        item.get("andrey_notes") or "",
    ]
    for col, val in enumerate(cells, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = THIN_BORDER

    # Hyperlink LinkedIn URL
    if item.get("linkedin"):
        link_cell = ws.cell(row=row, column=8)
        link_cell.hyperlink = item["linkedin"]
        link_cell.font = LINK_FONT
    else:
        ws.cell(row=row, column=8).fill = WARN_FILL

    # Pre-drafted message in mono font for readability
    msg_cell = ws.cell(row=row, column=11)
    msg_cell.font = MONO_FONT
    msg_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Chars vs limit — warn if over
    if chars > limit:
        ws.cell(row=row, column=12).fill = WARN_FILL
        ws.cell(row=row, column=12).font = Font(name="JetBrains Mono", size=9, bold=True, color="D44A4A")

    # Highlight rows that need Andrey's input
    if item["linkedin"] is None:
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col_idx).fill = WARN_FILL

    row += 1

# Column widths
widths = {
    "A": 4, "B": 8, "C": 11, "D": 22, "E": 32, "F": 22, "G": 18,
    "H": 38, "I": 50, "J": 18, "K": 70, "L": 7, "M": 7, "N": 12,
    "O": 14, "P": 13, "Q": 14, "R": 35,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Status dropdown for column O (Status)
status_dv = DataValidation(
    type="list",
    formula1='"Not started,In progress,Sent,Replied,Intro scheduled,Offer,Accepted,Declined,Skipped,Ghosted"',
    allow_blank=True,
)
status_dv.add(f"O2:O{row - 1}")
ws.add_data_validation(status_dv)

# Reply? Y/N
reply_dv = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
reply_dv.add(f"Q2:Q{row - 1}")
ws.add_data_validation(reply_dv)

# Conditional fill for Status=Sent (green) and Status=Skipped (gray)
green_fill = PatternFill("solid", fgColor="C9F2E1")
gray_fill = PatternFill("solid", fgColor="EEEEEE")
gold_fill = PatternFill("solid", fgColor="FFE9A8")

ws.conditional_formatting.add(
    f"O2:O{row - 1}",
    CellIsRule(operator="equal", formula=['"Sent"'], fill=green_fill),
)
ws.conditional_formatting.add(
    f"O2:O{row - 1}",
    CellIsRule(operator="equal", formula=['"Accepted"'], fill=gold_fill),
)
ws.conditional_formatting.add(
    f"O2:O{row - 1}",
    CellIsRule(operator="equal", formula=['"Skipped"'], fill=gray_fill),
)
ws.conditional_formatting.add(
    f"O2:O{row - 1}",
    CellIsRule(operator="equal", formula=['"Ghosted"'], fill=gray_fill),
)

# === SHEET 2: Quick Reference ===
qr = wb.create_sheet("Quick Reference")

qr["A1"] = "Quick Reference"
qr["A1"].font = Font(name="Inter", size=18, bold=True, color="1A2238")
qr.merge_cells("A1:E1")

qr["A3"] = "Evgeny's links"
qr["A3"].font = Font(name="Inter", size=12, bold=True, color="7C5CFF")
qr["A4"] = "LinkedIn"
qr["B4"] = LINKEDIN_EVGENY
qr["B4"].hyperlink = LINKEDIN_EVGENY
qr["B4"].font = LINK_FONT
qr["A5"] = "CV (PDF)"
qr["B5"] = "./assets/CV_Evgeny_Muravev.pdf (attach manually)"
qr["A6"] = "Email"
qr["B6"] = "evgeny.muravev@tabby.ai"
qr["A7"] = "Roles he's open to"
qr["B7"] = "Product roles in payments / lending. Open to relocation (London, Australia, US, Asia, Uzbekistan)."

qr["A9"] = "THE canonical product attribution (use verbatim)"
qr["A9"].font = Font(name="Inter", size=12, bold=True, color="7C5CFF")
qr["A10"] = "Evgeny helped us build our merchant lending and payments product, wallets for business in UAE"
qr["A10"].font = Font(name="JetBrains Mono", size=11, bold=True, italic=True, color="1A2238")
qr["A10"].fill = ACCENT_FILL
qr.merge_cells("A10:E10")
qr["A10"].alignment = Alignment(wrap_text=True, vertical="center")
qr.row_dimensions[10].height = 38

qr["A12"] = "The 6 ask types"
qr["A12"].font = Font(name="Inter", size=12, bold=True, color="7C5CFF")
ask_types = [
    ("send_cv", "Just forward CV — no role framing", "David Isakhanian, Evgeniya Kazina, Sarkis Atanesov"),
    ("recommend_for_role", "Refer to SPECIFIC job postings (job_urls in column I)", "Wise friend, Airwallex, Google, Uzum"),
    ("recommend_generally", "Recommend for any relevant role", "Fernando Plaza, Ricardo Saad, Oleg Dylevich, Oliver Hughes, Karun Kurien, Eugene Shilnikov, Michael Calvey, Harry Silvester, INB ex-CTO, Huspy VP"),
    ("ask_about_openings", "Ask if hiring in a specific area FIRST, then include job URLs", "Dr. Sebastian Walter (FAB)"),
    ("intro_to_founders", "Request an intro / make an intro to specific people", "Dwelly founders, BCG contacts, Yango regional"),
    ("post_in_chat", "Post Evgeny's CV in a group chat (not 1:1 DM)", "Veronika Nogovitsyna's UAE top managers chat"),
]
qr.cell(13, 1, "Ask type").fill = LABEL_FILL
qr.cell(13, 1).font = LABEL_FONT
qr.cell(13, 2, "What it means").fill = LABEL_FILL
qr.cell(13, 2).font = LABEL_FONT
qr.cell(13, 3, "Who it applies to").fill = LABEL_FILL
qr.cell(13, 3).font = LABEL_FONT
for i, (t, desc, who) in enumerate(ask_types, start=14):
    qr.cell(i, 1, t).font = MONO_FONT
    qr.cell(i, 2, desc).font = BODY_FONT
    qr.cell(i, 3, who).font = BODY_FONT
    qr.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
    qr.cell(i, 3).alignment = Alignment(wrap_text=True, vertical="top")

qr["A22"] = "Hard rules"
qr["A22"].font = Font(name="Inter", size=12, bold=True, color="7C5CFF")
rules = [
    "1. NEVER paraphrase the product attribution line. Use it verbatim.",
    "2. LinkedIn connection-request notes are HARD CAPPED at 300 characters.",
    "3. LinkedIn DMs are capped at 1900 characters (soft cap).",
    "4. Don't open with 'Hi, hope you're well at COMPANY' — open directly with the introduction.",
    "5. Always include job_urls when present — Evgeny specifically pointed at those roles.",
    "6. Always mention relocation when present — it tells the recipient he's serious.",
    "7. Sign off '— Andrey' on long-form, no sign-off on the 300-char connection note.",
    "8. NEVER auto-send. Andrey reviews and sends every message himself.",
]
for i, rule in enumerate(rules, start=23):
    qr.cell(i, 1, rule).font = BODY_FONT
    qr.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    qr.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")

qr.column_dimensions["A"].width = 20
qr.column_dimensions["B"].width = 60
qr.column_dimensions["C"].width = 55
qr.column_dimensions["D"].width = 20
qr.column_dimensions["E"].width = 20

# === SHEET 3: Needs Input ===
ni = wb.create_sheet("Needs Andrey Input")
ni["A1"] = "Items that need Andrey to identify the specific person before drafting"
ni["A1"].font = Font(name="Inter", size=14, bold=True, color="D44A4A")
ni.merge_cells("A1:F1")
ni["A2"] = "These 10 items have no LinkedIn URL because Evgeny didn't name a specific person. Andrey: please fill in column F with the name + LinkedIn URL, then move the row over to 'Outreach Plan'."
ni["A2"].font = Font(name="Inter", size=10, italic=True, color="8A93B0")
ni.merge_cells("A2:F2")
ni.row_dimensions[2].height = 40
ni["A2"].alignment = Alignment(wrap_text=True, vertical="top")

ni_headers = ["#", "ID", "Block", "Description", "Ask detail / context", "Andrey: fill in name + LinkedIn here"]
for col, h in enumerate(ni_headers, 1):
    c = ni.cell(4, col, h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = THIN_BORDER

ni_row = 5
for item in items_sorted:
    if item["linkedin"] is not None:
        continue
    cells = [
        item["priority"],
        item["id"],
        item["block"],
        item["name"],
        f"{item.get('ask_detail') or ''}\n\nAndrey notes: {item.get('andrey_notes') or '—'}",
        "",
    ]
    for col, val in enumerate(cells, 1):
        c = ni.cell(ni_row, col, val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = THIN_BORDER
    ni.cell(ni_row, 6).fill = WARN_FILL
    ni_row += 1

ni.column_dimensions["A"].width = 4
ni.column_dimensions["B"].width = 8
ni.column_dimensions["C"].width = 12
ni.column_dimensions["D"].width = 30
ni.column_dimensions["E"].width = 60
ni.column_dimensions["F"].width = 40

# === SHEET 4: Templates ===
tmpl = wb.create_sheet("Templates (reference)")
tmpl["A1"] = "Message templates per ask type"
tmpl["A1"].font = Font(name="Inter", size=14, bold=True, color="1A2238")
tmpl["A2"] = "These are the underlying templates. The 'Pre-drafted message' column in 'Outreach Plan' has these already filled in for each contact — but use these here for reference if you want to write a custom message."
tmpl["A2"].font = Font(name="Inter", size=10, italic=True, color="8A93B0")
tmpl.merge_cells("A2:C2")
tmpl["A2"].alignment = Alignment(wrap_text=True)
tmpl.row_dimensions[2].height = 40

template_rows = [
    ("recommend_generally", "LinkedIn DM (1st degree)",
     "Hi {NAME} — would like to introduce you to Evgeny who is exploring open roles at {COMPANY} in product. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE and I highly recommend him for any product role you may have in payments or lending.\n\nhis LinkedIn: linkedin.com/in/evgenymuravev\nhappy to share his CV directly.\n\n— Andrey"),
    ("recommend_generally", "Connection request note (≤300 chars)",
     "{NAME} — would like to introduce you to Evgeny who is exploring open roles at {COMPANY} in product. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE and I highly recommend him for any product role in payments or lending."),
    ("send_cv", "LinkedIn DM",
     "{NAME} — sending you the profile of Evgeny who is exploring new product roles. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE. Would love it if you could keep him in mind for any product roles in payments / lending at {COMPANY}.\n\nhis LinkedIn: linkedin.com/in/evgenymuravev\nCV attached.\n\n— Andrey"),
    ("recommend_for_role", "LinkedIn DM with job URLs",
     "{NAME} — would like to introduce you to Evgeny who is exploring open roles at {COMPANY} in product. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE and I highly recommend him.\n\nHe's specifically interested in these openings:\n- {JOB_URL_1}\n- {JOB_URL_2}\n\nHe's ready to relocate ({RELOCATION}).\n\nhis LinkedIn: linkedin.com/in/evgenymuravev\nhappy to share his CV directly.\n\n— Andrey"),
    ("ask_about_openings", "LinkedIn DM (two-part)",
     "{NAME} — hope you're doing well at {COMPANY}. Quick question: are you guys hiring on {AREA}? If yes, I'd like to introduce Evgeny who is exploring open roles at {COMPANY} in product. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE — highly recommend.\n\nHe's specifically interested in:\n- {JOB_URL_1}\n- {JOB_URL_2}\n\nhis LinkedIn: linkedin.com/in/evgenymuravev\nhappy to share his CV directly.\n\n— Andrey"),
    ("intro_to_founders", "LinkedIn DM",
     "Hi — would like to introduce you to Evgeny who is exploring new product opportunities. Evgeny helped us build our merchant lending and payments product, wallets for business in UAE.\n\n{CONTEXT}\n\nIf you're open to a 20-min chat, I'll connect you both. His LinkedIn: linkedin.com/in/evgenymuravev\n\n— Andrey"),
    ("post_in_chat", "WhatsApp / Telegram chat post",
     "sharing the profile of Evgeny Muravev — he helped us build Tabby's merchant lending and payments product, wallets for business in UAE. Exploring new product roles in payments / lending, UAE based, open to relocation.\n\nLinkedIn: linkedin.com/in/evgenymuravev\nCV attached.\n\nhappy to make any intro — DM me directly.\n\n— Andrey"),
]

tmpl_h = ["Ask type", "Channel / variant", "Template"]
for col, h in enumerate(tmpl_h, 1):
    c = tmpl.cell(4, col, h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = THIN_BORDER

for i, (ask, variant, body) in enumerate(template_rows, start=5):
    tmpl.cell(i, 1, ask).font = MONO_FONT
    tmpl.cell(i, 2, variant).font = LABEL_FONT
    tmpl.cell(i, 3, body).font = MONO_FONT
    for col in range(1, 4):
        tmpl.cell(i, col).alignment = Alignment(wrap_text=True, vertical="top")
        tmpl.cell(i, col).border = THIN_BORDER
    tmpl.row_dimensions[i].height = 130

tmpl.column_dimensions["A"].width = 22
tmpl.column_dimensions["B"].width = 28
tmpl.column_dimensions["C"].width = 100

# === Save and report ===
out = HERE / "andrey_outreach_sheet.xlsx"
wb.save(out)

# Report stats
total = len(items_sorted)
needs_input = sum(1 for x in items_sorted if x["linkedin"] is None)
over_limit = 0
for item in items_sorted:
    msg = generate_message(item)
    if len(msg) > char_limit(item):
        over_limit += 1
        print(f"  ⚠  {item['name']} message is {len(msg)} chars vs limit {char_limit(item)}")

print(f"\n✓ Wrote {out}")
print(f"  - {total} contacts on 'Outreach Plan' tab")
print(f"  - {needs_input} contacts on 'Needs Andrey Input' tab (no LinkedIn URL yet)")
print(f"  - 7 templates on 'Templates (reference)' tab")
print(f"  - Quick Reference tab with rules + canonical product line")
if over_limit:
    print(f"  ⚠  {over_limit} messages exceed their char limit — trim before sending")
